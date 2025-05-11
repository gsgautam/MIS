import os
from dotenv import load_dotenv

load_dotenv() # Loads variables from .env into environment variables

# --- MySQL Configuration ---
MYSQL_CONFIG = {
    'user': os.getenv('DB_USER'),
    'password': os.getenv('DB_PASSWORD'),
    'host': os.getenv('DB_HOST'),
    'database': os.getenv('DB_NAME'),
    'raise_on_warnings': True
}

# --- Admin Sign Up Secret Code ---
ADMIN_SIGNUP_SECRET = os.getenv('ADMIN_SECRET')

# --- Email Configuration ---
EMAIL_ACCOUNT = os.getenv('EMAIL_USER')
EMAIL_PASSWORD = os.getenv('EMAIL_PASS')
IMAP_SERVER = os.getenv('IMAP_SVR')

import os
import sys # Make sure sys is imported if not already
import json # Added for handling JSON data for file lists

# --- Configuration Loading ---
# These will be global and available to the rest of the script.
MYSQL_CONFIG = None
ADMIN_SIGNUP_SECRET = None
EMAIL_ACCOUNT = None
EMAIL_PASSWORD = None
IMAP_SERVER = None

try:
    from instance import config
    MYSQL_CONFIG = config.MYSQL_CONFIG
    ADMIN_SIGNUP_SECRET = config.ADMIN_SIGNUP_SECRET
    EMAIL_ACCOUNT = config.EMAIL_ACCOUNT
    EMAIL_PASSWORD = config.EMAIL_PASSWORD
    IMAP_SERVER = config.IMAP_SERVER
    # You could also load APP_NAME from config.py if you want it configurable:
    # APP_NAME = config.APP_NAME
    print("Successfully loaded configuration from instance/config.py")
except ImportError:
    error_message = "CRITICAL ERROR: instance/config.py not found. Please create it in the 'instance' folder with your credentials and restart."
    print(error_message)
    # Attempt to show error in Streamlit if it's already loaded
    if 'streamlit' in sys.modules:
        st_module = sys.modules['streamlit']
        if hasattr(st_module, 'error') and hasattr(st_module, 'stop'):
            st_module.error(error_message)
            st_module.stop()
    sys.exit(error_message) # Exit in any case
except AttributeError as e:
    error_message = f"CRITICAL ERROR: A configuration variable ({e}) is missing from instance/config.py. Please add it and restart."
    print(error_message)
    if 'streamlit' in sys.modules:
        st_module = sys.modules['streamlit']
        if hasattr(st_module, 'error') and hasattr(st_module, 'stop'):
            st_module.error(error_message)
            st_module.stop()
    sys.exit(error_message) # Exit in any case
# --- End Configuration Loading ---

# --- Global App Constants (Non-Secrets) ---
# Define these after config loading, so they are globally available.
APP_NAME = "Office MIS (MySQL)"  # Or load from config.py if you added it there
APP_ROOT_PATH = os.path.dirname(os.path.abspath(__file__))

# --- Main Script Imports ---
# sys and os are already imported at the top.
import subprocess
import calendar
import pandas as pd
import numpy as np # Import numpy for np.datetime64
try:
    PANDAS_AVAILABLE = True
    import openpyxl # For custom excel
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    PANDAS_AVAILABLE = False
    print("Pandas or Openpyxl not available. Some features might be disabled.")

import mysql.connector
from mysql.connector import Error # Error is now correctly imported
import datetime # This was imported in the original, ensure it's here.
from io import BytesIO

# --- Passlib Import and Initialization with Error Handling ---
try:
    from passlib.context import CryptContext
    pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
    PASSLIB_AVAILABLE = True
    print("Passlib loaded successfully.")
except AttributeError as e_passlib: # bcrypt version issue
    print("="*60)
    print("ERROR: Potential incompatibility detected between 'passlib' and 'bcrypt'.")
    print(f"Details: {e_passlib}")
    print("RECOMMENDATION: Please try reinstalling these libraries:")
    print("  pip uninstall bcrypt passlib")
    print("  pip install bcrypt passlib")
    print("Password hashing functionality will be disabled.")
    print("="*60)
    PASSLIB_AVAILABLE = False
    pwd_context = None
except ImportError:
    print("ERROR: 'passlib' library not found. Password hashing disabled. Install with 'pip install passlib[bcrypt]'")
    PASSLIB_AVAILABLE = False
    pwd_context = None
# --- End Passlib Import ---

import imaplib
import email
from email.header import decode_header
import re
import streamlit as st # Import Streamlit

# --- Self-Launch Helper Function ---
def run_with_streamlit():
    print("Attempting to relaunch with Streamlit...")
    try:
        python_executable = sys.executable
        script_path = os.path.abspath(__file__)
        command = [python_executable, "-m", "streamlit", "run", script_path, "--", "--launched-by-script"]
        print(f"Executing: {' '.join(command)}")
        subprocess.Popen(command)
        sys.exit(0)
    except Exception as e:
        print(f"Error relaunching: {e}\nPlease run manually: streamlit run \"{script_path}\"")
        input("Press Enter to exit...")
        sys.exit(1)

# --- Main Execution Control ---
if __name__ == "__main__":
    NEEDS_RELAUNCH = (len(sys.argv) <= 1 or sys.argv[-1] != '--launched-by-script')

    if NEEDS_RELAUNCH:
        print("Script started directly. Relaunching via Streamlit...")
        try:
            import streamlit # Test if installed, st is not yet defined here
        except ImportError:
            print("STOP: Streamlit not installed ('pip install streamlit')")
            input("Press Enter to exit...")
            sys.exit(1)
        run_with_streamlit() # Calls sys.exit()
    else:
        # --- Streamlit Application Code Execution ---
        # This block runs when executed by 'streamlit run ...' or after relaunch
        print("Streamlit application starting...")

        # APP_NAME and APP_ROOT_PATH are now defined globally at the top.
        # Secret configurations (MYSQL_CONFIG etc.) are also loaded globally.

        # --- Page Config ---
        # Uses the globally defined APP_NAME
        st.set_page_config(page_title=APP_NAME, layout="wide", initial_sidebar_state="expanded")

        # --- Bank Names List (Deduplicated and Categorized) ---
        # These are app-specific constants, not secrets.
        BANKS_PORTAL_REPORT = sorted([
            "Aditya Birla HL (Portal Report)", "Chola Mandalam (Micro LAP) (Portal Report)",
            "Chola Mandalam (Prime LAP) (Portal Report)", "CSL (Portal Report)",
            "CSC (Portal Report)", "DMI (Portal Report)", "ICICI (Portal Report)",
            "IIFL (Portal Report)", "Kotak (Portal Report)", "Motilal (Portal Report)",
            "Piramal (Portal Report)", "Shubham HFC (Portal Report)", "TATA (Portal Report)"
        ])

        BANKS_NORMAL = sorted(list(set([ # Using set to ensure uniqueness from various sources
            "AU Small Finance Bank", "Ambit Finance", "Aye Finance", "Axis", "Canara bank",
            "Chola (HL)", "Chola (SME)", "Chola Mandalam (SBPL)", "DCB Bank",
            "Grihum", "Godrej Finance Ltd", "HDFC Bank", "Hero Housing", "ICICI HFC",
            "IDFC", "Incred", "IndusInd", "Jana bank", "L&T", "LICHFL", "Mahindra",
            "Poonawalla", "SK Finance", "SMFG", "True home",
            "Utkarsh Small Finance Bank", "Ujjivan Small Finance Bank", "Yes Bank"
        ])))
        ALL_BANK_OPTIONS_COMBINED = sorted(list(set(BANKS_PORTAL_REPORT + BANKS_NORMAL)))
        ALL_BANK_OPTIONS_DROPDOWN = ["--Select Bank--"] + ALL_BANK_OPTIONS_COMBINED + ["Other"]
        ALL_BANK_OPTIONS_FILTER = ["-- All Banks --"] + ALL_BANK_OPTIONS_COMBINED

        # === UTILITY FUNCTIONS ===
        def verify_password(plain, hashed):
            if not PASSLIB_AVAILABLE:
                st.error("Password verification disabled due to library error.")
                return False
            if not plain or not hashed: return False
            try: return pwd_context.verify(plain, hashed)
            except Exception as e: print(f"Verify Password Error: {e}"); return False

        def get_password_hash(pwd):
            if not PASSLIB_AVAILABLE:
                st.error("Password hashing disabled due to library error.")
                return None
            return pwd_context.hash(pwd)

        @st.cache_resource
        def initialize_database():
            print("Checking DB connection and table existence...")
            if MYSQL_CONFIG is None: # Check if config loading failed
                print("FATAL: MYSQL_CONFIG is not loaded. Cannot initialize database.")
                return False
            try:
                conn = mysql.connector.connect(**MYSQL_CONFIG)
                cursor = conn.cursor()
                tables = ['admins', 'users', 'office', 'site_engineers']
                exists = True
                for t in tables:
                    cursor.execute(f"SHOW TABLES LIKE '{t}';")
                    if not cursor.fetchone():
                        print(f"FATAL: Table '{t}' not found.")
                        exists = False
                cursor.close()
                conn.close()
                if exists:
                    print("DB connection & tables OK.")
                # IMPORTANT: Inform user about new columns needed for file uploads
                # This is a conceptual check, actual check might be more involved or done at app startup
                if exists:
                    conn_check_cols = mysql.connector.connect(**MYSQL_CONFIG)
                    cursor_check_cols = conn_check_cols.cursor()
                    cursor_check_cols.execute("SHOW COLUMNS FROM office LIKE 'site_photo_filenames';")
                    if not cursor_check_cols.fetchone():
                        print("WARNING: Column 'site_photo_filenames' not found in 'office' table. File upload feature may not work correctly.")
                        # st.warning("Database column 'site_photo_filenames' is missing. File features might be affected.")
                    cursor_check_cols.execute("SHOW COLUMNS FROM office LIKE 'site_document_filenames';")
                    if not cursor_check_cols.fetchone():
                        print("WARNING: Column 'site_document_filenames' not found in 'office' table. File upload feature may not work correctly.")
                        # st.warning("Database column 'site_document_filenames' is missing. File features might be affected.")
                    cursor_check_cols.close()
                    conn_check_cols.close()

                return exists
            except Error as e: # mysql.connector.Error
                print(f"MySQL Init Error: {e}")
                return False
            except Exception as e_gen: # Catch other potential errors like config issues if not caught earlier
                print(f"General Error during DB Init: {e_gen}")
                return False


        def run_db_query(query, params=(), fetch_one=False, fetch_all=False):
            conn = None
            cursor = None
            results = None
            if MYSQL_CONFIG is None:
                st.error("Database configuration not loaded. Cannot run query.")
                print("DB Error: MYSQL_CONFIG is None.")
                return None
            try:
                conn = mysql.connector.connect(**MYSQL_CONFIG)
                cursor = conn.cursor(dictionary=True)
                cursor.execute(query, params)
                if fetch_one:
                    results = cursor.fetchone()
                elif fetch_all:
                    results = cursor.fetchall()
                else:
                    conn.commit()
                    results = cursor.lastrowid
            except Error as e: # mysql.connector.Error
                st.error(f"Database Error. See console log.")
                print(f"DB Error: {e} | Query: {query} | Params: {params}")
                results = None # Ensure results is None on error
                if conn and conn.is_connected() and not fetch_one and not fetch_all:
                    try:
                        conn.rollback()
                        print("Transaction rolled back.")
                    except Error as rb_e: # mysql.connector.Error
                        print(f"Rollback failed: {rb_e}")
            finally:
                if cursor:
                    cursor.close()
                if conn and conn.is_connected():
                    conn.close()
            return results

        def add_lead_to_db(lead_data):
            lead_data.setdefault('received_date', datetime.datetime.now())
            lead_data.setdefault('status', 'New')
            if not lead_data.get('bank_name') or lead_data['bank_name'] == "--Select Bank--":
                st.error("Bank Name required.")
                return False
            if not lead_data.get('property_details'):
                st.error("Property Details required.")
                return False
            expected_columns = [
                'bank_name', 'property_details', 'received_date', 'deadline', 'status',
                'site_engineer', 'report_creator', 'report_issue_notes', 'admin_review_status', 'admin_comments',
                'date_of_allocation', 'customer_name', 'application_number', 'location', 'contact_number',
                'site_link', 'visit_initiation_date', 'visit_completion_date', 'lead_completion_date',
                'appraiser_quotation_obs', 'distance', 'visit_type', 'remarks', 'branch_virtual',
                'site_photo_filenames', 'site_document_filenames' # Added new columns
            ]
            columns_to_insert = [col for col in expected_columns if col in lead_data and lead_data[col] is not None]
            if 'received_date' not in columns_to_insert: columns_to_insert.append('received_date')
            if 'status' not in columns_to_insert: columns_to_insert.append('status')

            params_list = []
            for col in columns_to_insert:
                value = lead_data.get(col)
                if value is None: param_val = None
                elif isinstance(value, datetime.datetime): param_val = value.strftime('%Y-%m-%d %H:%M:%S')
                elif isinstance(value, datetime.date): param_val = value.strftime('%Y-%m-%d')
                elif PANDAS_AVAILABLE and pd.isna(value): param_val = None
                else: param_val = value
                params_list.append(param_val)

            if not columns_to_insert:
                st.error("No data to insert.")
                return False
            query = f"INSERT INTO office ({', '.join(['`'+col+'`' for col in columns_to_insert])}) VALUES ({', '.join(['%s'] * len(params_list))})"
            params_tuple = tuple(params_list) # Use a different variable name

            insert_id = run_db_query(query, params_tuple) # Pass the tuple
            success = insert_id is not None and insert_id > 0
            if success:
                print(f"DB Insert OK, ID: {insert_id}")
            else:
                print(f"DB Insert FAILED. Query generated: {query} | Params used: {params_tuple}")
            return success

        def get_email_body(msg):
            body = None; charset = None
            if msg.is_multipart():
                for part in msg.walk():
                    ctype = part.get_content_type(); cdispo = str(part.get('Content-Disposition'))
                    if ctype == 'text/plain' and 'attachment' not in cdispo:
                        charset = part.get_content_charset()
                        try: body = part.get_payload(decode=True).decode(charset or 'utf-8', errors='ignore'); break
                        except Exception as e: print(f"Error decoding multipart: {e}")
            else:
                ctype = msg.get_content_type()
                if ctype == 'text/plain':
                    charset = msg.get_content_charset()
                    try: body = msg.get_payload(decode=True).decode(charset or 'utf-8', errors='ignore')
                    except Exception as e: print(f"Error decoding single part: {e}")
            return body

        def parse_subject(subject_header_val):
            subject = "No Subject"
            if subject_header_val:
                try:
                    parts = []; decoded_header = decode_header(subject_header_val)
                    for part_content,charset in decoded_header:
                        if isinstance(part_content, bytes): parts.append(part_content.decode(charset or 'utf-8', 'ignore'))
                        else: parts.append(part_content)
                    subject = "".join(parts)
                except Exception as e: print(f"Error decoding subject: {e}"); subject = str(subject_header_val)
            return subject

        def extract_info_from_email(subject_text, body_text, sender_email):
            print(f"\n--- Parsing Email --- From: {sender_email}, Subject: {subject_text}")
            extracted = {}
            if not body_text: body_text = "" # Ensure body_text is a string
            try:
                match_prop = re.search(r"Property(?: Address| Details| Location):\s*(.*?)(?:\n\n|Due Date:|Deadline:|$)", body_text, re.IGNORECASE | re.DOTALL)
                if match_prop: extracted['property_details'] = match_prop.group(1).strip().replace('\r\n', ' ').replace('\n', ' ')[:400]
                else: extracted['property_details'] = body_text[:200].strip().replace('\r\n', ' ').replace('\n', ' ')

                date_kw = [r"Due Date", r"Deadline", r"Valuation Required By", r"Submit By"]; date_pats = [r"(\d{1,2}[-/]\d{1,2}[-/]\d{2,4})", r"(\d{4}[-/]\d{1,2}[-/]\d{1,2})"]; deadline_str = None
                for kw in date_kw:
                    if deadline_str: break
                    for pat in date_pats:
                        match_d = re.search(rf"{kw}[:\s]*{pat}", body_text, re.IGNORECASE)
                        if match_d: deadline_str = match_d.group(1).strip(); break
                if deadline_str:
                    parsed_dt = None; fmts = ["%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d", "%d/%m/%y", "%d-%m/%y", "%m/%d/%Y", "%m-%d-%Y", "%Y/%m/%d", "%m/%d/%y", "%m-%d/%y"] # Added YY/MM/DD
                    for fmt in fmts:
                        try: parsed_dt = datetime.datetime.strptime(deadline_str, fmt); extracted['deadline'] = parsed_dt.strftime('%Y-%m-%d'); break
                        except ValueError: continue
                    if not parsed_dt: print(f"Could not parse deadline: {deadline_str}")

                std_bank = None
                for known_bank_option in ALL_BANK_OPTIONS_COMBINED: # Ensure this list is defined
                    base_name = known_bank_option.split(" (")[0].strip().lower()
                    if re.search(r'\b' + re.escape(base_name) + r'\b', sender_email.lower()) or \
                       re.search(r'\b' + re.escape(base_name) + r'\b', subject_text.lower()) or \
                       re.search(r'\b' + re.escape(base_name) + r'\b', body_text.lower()): # Added body_text search for bank
                        std_bank = known_bank_option
                        break

                if std_bank: extracted['bank_name'] = std_bank
                else:
                    match_bank_body = re.search(r"Bank Name[:\s]+(.*?)(?:\n|$)", body_text, re.IGNORECASE)
                    if match_bank_body: extracted['bank_name'] = match_bank_body.group(1).strip()
                    else: extracted['bank_name'] = sender_email.split('@')[0] if '@' in sender_email else sender_email

                match_cust = re.search(r"(?:Customer|Client) Name[:\s]+(.*?)(?:\n|$)", body_text, re.IGNORECASE)
                if match_cust: extracted['customer_name'] = match_cust.group(1).strip()
                match_app = re.search(r"Application (?:Number|No|ID)[:\s]+([\w-]+)", body_text, re.IGNORECASE)
                if match_app: extracted['application_number'] = match_app.group(1).strip()
                match_loc = re.search(r"Location[:\s]+(.*?)(?:\n\n|$)", body_text, re.IGNORECASE | re.DOTALL) # Ensure it's DOTALL
                if match_loc: extracted['location'] = match_loc.group(1).strip().replace('\r\n', ' ').replace('\n', ' ')
                else: extracted['location'] = extracted.get('property_details') # Fallback
                match_contact = re.search(r"(?:Contact|Phone|Mobile) (?:Number|No)[:\s]+([\d\s()-+]+)", body_text, re.IGNORECASE)
                if match_contact: extracted['contact_number'] = match_contact.group(1).strip()

            except Exception as e_extract: print(f"ERROR during email info extraction: {e_extract}"); return None
            if not extracted.get('property_details') or not extracted.get('bank_name'): print(f"Core details missing: Subject='{subject_text}'"); return None
            print(f"Extraction result (Email): {extracted}"); return extracted

        def check_emails_once():
            st.info("Checking emails... Please wait.")
            print(f"\n[{datetime.datetime.now()}] == Starting Email Check ==")
            added_count = 0; processed_ids_list = []; mail = None
            # Ensure EMAIL_ACCOUNT, EMAIL_PASSWORD, IMAP_SERVER are loaded from config
            if not EMAIL_ACCOUNT or EMAIL_ACCOUNT == "your_mis_email@gmail.com" or \
               not EMAIL_PASSWORD or EMAIL_PASSWORD == "YOUR_APP_PASSWORD" or \
               not IMAP_SERVER:
                st.error("Email credentials or IMAP server are not configured properly in instance/config.py!")
                print("Email config error: Credentials or server not set.")
                return 0
            try:
                mail = imaplib.IMAP4_SSL(IMAP_SERVER)
                mail.login(EMAIL_ACCOUNT, EMAIL_PASSWORD)
                mail.select("inbox")
                print("Connected. Searching for unseen emails...")
                status, messages_bytes = mail.search(None, "(UNSEEN)")
                if status != "OK":
                    st.warning(f"Failed to search emails: {status}")
                    print(f"Email search failed: {status}")
                    return 0

                email_id_list_bytes = messages_bytes[0].split()
                if not email_id_list_bytes or (len(email_id_list_bytes) == 1 and not email_id_list_bytes[0]):
                    st.info("No new unseen emails found.")
                    print("No unseen emails.")
                    return 0

                st.info(f"Found {len(email_id_list_bytes)} unseen emails.")
                for email_id_b in email_id_list_bytes:
                    email_id_s = email_id_b.decode()
                    print(f"Processing Email ID: {email_id_s}")
                    try:
                        res, msg_data_list = mail.fetch(email_id_b, "(RFC822)")
                        if res == "OK":
                            for response_part in msg_data_list:
                                if isinstance(response_part, tuple) and len(response_part) >= 2 and isinstance(response_part[1], bytes):
                                    try:
                                        msg_obj = email.message_from_bytes(response_part[1])
                                        subject = parse_subject(msg_obj["Subject"])
                                        sender = email.utils.parseaddr(msg_obj.get('From'))[1]
                                        body = get_email_body(msg_obj)
                                        if body:
                                            lead_info = extract_info_from_email(subject, body, sender)
                                            if lead_info and add_lead_to_db(lead_info):
                                                print(f"  Success: Added lead from Email ID {email_id_s} to DB.")
                                                added_count += 1
                                                processed_ids_list.append(email_id_b)
                                            elif lead_info: print(f"  FAIL: Add lead from Email ID {email_id_s} to DB failed.")
                                            else: print(f"  Info: Could not extract lead info from Email ID {email_id_s}.")
                                        else: print(f"  Info: No text body for Email ID {email_id_s}.")
                                    except Exception as e_parse: print(f"  ERROR parsing content of Email ID {email_id_s}: {e_parse}")
                        else: print(f"Failed to fetch Email ID: {email_id_s}. Status: {res}")
                    except Exception as e_fetch_process: print(f"  ERROR fetching/processing Email ID {email_id_s}: {e_fetch_process}")

                if processed_ids_list:
                    ids_to_mark_bytes = b','.join(processed_ids_list)
                    print(f"Marking {len(processed_ids_list)} emails as read...")
                    store_status, _ = mail.store(ids_to_mark_bytes, '+FLAGS', '\\Seen')
                    if store_status == 'OK': print("  Successfully marked emails as read.")
                    else: print(f"  Failed to mark emails as read: {store_status}")
            except imaplib.IMAP4.error as e_imap:
                st.error(f"IMAP Error: {e_imap}. Check credentials and IMAP server settings.")
                print(f"IMAP Error: {e_imap}")
            except Exception as e_check:
                st.error(f"Unexpected error during email check: {e_check}")
                print(f"Unexpected error during email check: {e_check}")
            finally:
                if mail:
                    try:
                        if mail.state != 'LOGOUT': # Check if not already logged out
                           mail.close()
                           mail.logout()
                           print("Email connection closed.")
                    except Exception as e_logout: print(f"Error during email final logout/close: {e_logout}")

            end_message = f"Email check finished. New leads added: {added_count}"
            if added_count > 0: st.success(end_message)
            else: st.info(end_message)
            return added_count

        def get_status_and_color_value(item):
            status = item.get('status', 'New'); deadline = item.get('deadline'); creator = item.get('report_creator'); d_str = None
            if isinstance(deadline, (datetime.date, datetime.datetime)): d_str = deadline.strftime('%Y-%m-%d')
            elif isinstance(deadline, str):
                try: datetime.datetime.strptime(deadline, '%Y-%m-%d'); d_str = deadline
                except ValueError: d_str = None

            clrs={'normal':'white','red':'#ffdddd','orange':'#ffe8cc','green':'#ddffdd','grey':'#e0e0e0','blue':'#ddeeff','purple':'#e8ddff','yellow_assign':'#ffffcc','yellow_report':'#fffacd'}
            key = 'normal'; final_status = str(status)

            if status == 'Completed': key = 'green'
            elif status == 'Visit Done' and not creator: key = 'yellow_report'; final_status += " (Pend Report)"
            elif status == 'Report in Progress': key = 'purple'
            elif status == 'Visit Done': key = 'blue'
            elif status == 'Assigned Engineer': key = 'yellow_assign'
            elif status == 'On Hold': key = 'orange'

            if d_str and status != 'Completed':
                try:
                    deadline_date = datetime.datetime.strptime(d_str, '%Y-%m-%d').date()
                    days_to_deadline = (deadline_date - datetime.date.today()).days
                    if days_to_deadline < 0:
                        key = 'grey'
                        final_status = final_status.split(" (")[0] + " (Overdue)"
                    elif days_to_deadline <= 2: key = 'red'
                    elif days_to_deadline <= 5 and key not in ['red', 'grey']: key = 'orange'
                except Exception as e: print(f"Deadline color calculation error: {e}")
            return final_status, clrs.get(key, 'white')

        # --- Dashboard Functions ---
        def display_summary_dashboard_stats(df):
            st.subheader("Overall MIS Summary")
            if not PANDAS_AVAILABLE or df is None or df.empty: st.info("No data available for summary."); return
            if 'site_engineer' not in df.columns or 'status' not in df.columns: st.warning("Required columns ('site_engineer', 'status') missing for summary."); return

            vp_df = df[df['status'] == 'Assigned Engineer']['site_engineer'].value_counts().reset_index(); vp_df.columns = ['Engineer', 'VP Count']; vp_df = vp_df[vp_df['Engineer'].notna() & (vp_df['Engineer'] != '')]
            vd_sts = ['Visit Done', 'Report in Progress', 'Completed']; vd_df = df[df['status'].isin(vd_sts)]['site_engineer'].value_counts().reset_index(); vd_df.columns = ['Engineer', 'VD Count']; vd_df = vd_df[vd_df['Engineer'].notna() & (vd_df['Engineer'] != '')]

            c1, c2 = st.columns(2)
            with c1: st.markdown("##### Visit Pending (VP) by Engineer"); st.dataframe(vp_df if not vp_df.empty else pd.DataFrame(columns=['Engineer', 'VP Count']),use_container_width=True, hide_index=True)
            with c2: st.markdown("##### Visit Done (VD) by Engineer"); st.dataframe(vd_df if not vd_df.empty else pd.DataFrame(columns=['Engineer', 'VD Count']), use_container_width=True, hide_index=True)

            st.markdown("---")
            issue = {}
            issue['Rejected/Revision'] = df[df.get('admin_review_status', pd.Series(dtype=str)) == 'Rejected - Needs Revision'].shape[0]
            issue['On Hold'] = df[df.get('status', pd.Series(dtype=str)) == 'On Hold'].shape[0]
            issue['New (Unassigned)'] = df[df.get('status', pd.Series(dtype=str)) == 'New'].shape[0]
            st.markdown("##### Key Issues Summary"); issue_df = pd.DataFrame(list(issue.items()), columns=['Category', 'Count']); st.dataframe(issue_df, use_container_width=True, hide_index=True); st.markdown("---")

            st.markdown("##### Overall Status Counts")
            total_leads = df.shape[0]
            visit_done_count = df[df['status'].isin(['Visit Done', 'Report in Progress', 'Completed'])].shape[0]
            visit_pending_count = df[df['status'] == 'Assigned Engineer'].shape[0]
            completed_count = df[df['status'] == 'Completed'].shape[0]
            report_in_progress_count = df[df['status'] == 'Report in Progress'].shape[0]
            pending_delivery_count = df[
                (df['status'] == 'Report in Progress') |
                ((df['status'] == 'Completed') & (df.get('admin_review_status', pd.Series(dtype=str)) == 'Pending Review'))
            ].shape[0]

            summary_counts = {
                "Total Leads": total_leads,
                "Visits Actually Done": visit_done_count,
                "Visits Pending (Assigned)": visit_pending_count,
                "Reports Completed": completed_count,
                "Reports in Progress": report_in_progress_count,
                "Pending Admin Review/Delivery": pending_delivery_count
            }
            summary_df = pd.DataFrame.from_dict(summary_counts, orient='index', columns=['Count']); st.dataframe(summary_df, use_container_width=True); st.markdown("---")

        def display_per_day_allocation_dashboard(df, yr, mo):
            st.subheader("PER DAY TOTAL VISIT ALLOCATION TO SITE ENGINEERS")
            month_name = calendar.month_name[mo]
            try:
                _, num_days_in_month = calendar.monthrange(yr, mo)
                start_date_str = f"01 {month_name.upper()}, {yr}"
                end_date_str = f"{num_days_in_month} {month_name.upper()}, {yr}"
                st.write(f"Report for: ({start_date_str} TO {end_date_str})")
            except ValueError:
                st.error(f"Invalid year/month selected: {yr}/{mo}")
                return

            if not PANDAS_AVAILABLE: st.warning("Pandas library needed for this report."); return
            if df is None or df.empty: st.info(f"No lead data found for {month_name} {yr}."); return
            if 'received_date' not in df.columns: st.error("Report requires 'received_date' column in the data."); return

            try:
                daily_df = df.copy()
                # Ensure 'received_date' is datetime before accessing .dt
                daily_df['recv_date_dt'] = pd.to_datetime(daily_df['received_date'], errors='coerce')
                # Remove timezone if present, before accessing .dt.date
                if pd.api.types.is_datetime64_any_dtype(daily_df['recv_date_dt']) and daily_df['recv_date_dt'].dt.tz is not None:
                    daily_df['recv_date_dt'] = daily_df['recv_date_dt'].dt.tz_localize(None)
                daily_df['recv_date'] = daily_df['recv_date_dt'].dt.date
                daily_df.dropna(subset=['recv_date'], inplace=True) # Drop rows where date conversion failed

                month_df = daily_df[(daily_df['recv_date_dt'].dt.year == yr) & (daily_df['recv_date_dt'].dt.month == mo)]
            except Exception as e: st.error(f"Error processing dates for daily report: {e}"); return

            if month_df.empty: st.info(f"No leads received in {month_name} {yr}."); return

            site_engineers_list = []
            if 'site_engineer' in month_df.columns:
                assigned_in_month_df = month_df[month_df['status'] == 'Assigned Engineer']
                if not assigned_in_month_df.empty:
                    site_engineers_list = sorted(assigned_in_month_df['site_engineer'].dropna().unique().tolist())
                    site_engineers_list = [eng for eng in site_engineers_list if str(eng).strip()]

            days_in_month_dates = [datetime.date(yr, mo, d) for d in range(1, num_days_in_month + 1)]
            report_data_list = []

            base_cols = ["Date", "Total Received", "Completed This Day", "Visit Done This Day", "Visit Pending (Assigned)"]
            dynamic_engineer_cols = site_engineers_list
            status_cols = ["New (Mail)", "On Hold", "Rejected/Revision"]
            report_columns_final_order = base_cols + dynamic_engineer_cols + status_cols

            for specific_day in days_in_month_dates:
                leads_on_specific_day = month_df[month_df['recv_date'] == specific_day]
                row_data_dict = {'Date': specific_day.strftime('%d %B %Y').upper()}

                row_data_dict['Total Received'] = leads_on_specific_day.shape[0]
                row_data_dict['Completed This Day'] = leads_on_specific_day[leads_on_specific_day['status'] == 'Completed'].shape[0]
                row_data_dict['Visit Done This Day'] = leads_on_specific_day[leads_on_specific_day['status'].isin(['Visit Done', 'Report in Progress'])].shape[0]
                row_data_dict['Visit Pending (Assigned)'] = leads_on_specific_day[leads_on_specific_day['status'] == 'Assigned Engineer'].shape[0]

                for eng_name in site_engineers_list:
                    row_data_dict[eng_name] = leads_on_specific_day[
                        (leads_on_specific_day['site_engineer'] == eng_name) &
                        (leads_on_specific_day['status'] == 'Assigned Engineer')
                    ].shape[0]

                row_data_dict['New (Mail)'] = leads_on_specific_day[leads_on_specific_day.get('status', pd.Series(dtype=str)) == 'New'].shape[0]
                row_data_dict['On Hold'] = leads_on_specific_day[leads_on_specific_day.get('status', pd.Series(dtype=str)) == 'On Hold'].shape[0]
                row_data_dict['Rejected/Revision'] = leads_on_specific_day[leads_on_specific_day.get('admin_review_status', pd.Series(dtype=str)) == 'Rejected - Needs Revision'].shape[0]

                report_data_list.append(row_data_dict)

            if report_data_list:
                display_df = pd.DataFrame(report_data_list, columns=report_columns_final_order)
                display_df.fillna(0, inplace=True)
                for col_name in report_columns_final_order:
                    if col_name != 'Date':
                        display_df[col_name] = display_df[col_name].astype(int)
                st.dataframe(display_df, use_container_width=True, height=min(len(report_data_list) * 35 + 38, 600), hide_index=True)
            else: st.info(f"No processed lead data available to display for {month_name} {yr}.")
            st.markdown("---")

        def generate_custom_excel_report(df_all_leads):
            if not PANDAS_AVAILABLE:
                st.error("Pandas and Openpyxl are required for this Excel report.")
                return None

            visit_done_statuses = ['Visit Done', 'Report in Progress', 'Completed']
            visit_pending_statuses = ['Assigned Engineer', 'New', 'On Hold']

            df_visit_done = df_all_leads[df_all_leads['status'].isin(visit_done_statuses)].copy()
            df_visit_pending = df_all_leads[df_all_leads['status'].isin(visit_pending_statuses)].copy()

            for df_temp in [df_visit_done, df_visit_pending]:
                date_cols_to_format = ['date_of_allocation', 'received_date', 'deadline', 'visit_initiation_date', 'visit_completion_date', 'lead_completion_date']
                for col in date_cols_to_format:
                    if col in df_temp.columns:
                        df_temp.loc[:, col] = pd.to_datetime(df_temp[col], errors='coerce').dt.date


            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "MIS Report"

            header_font = Font(name='Calibri', size=11, bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            cell_font = Font(name='Calibri', size=11)
            center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            border_thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

            current_date_str = datetime.date.today().strftime("%d-%b-%Y")
            report_columns = ["Sr. No.", "Bank Name", "Branch/Virtual", "Received Date", "Customer's Name", "Location", "Site Engg.", "Deadline", "Status"]
            current_row = 1

            # VISIT DONE Section
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(report_columns))
            title_cell_done = ws.cell(row=current_row, column=1, value=f"VISIT DONE - as on {current_date_str}")
            title_cell_done.font = Font(name='Calibri', size=14, bold=True)
            title_cell_done.alignment = center_alignment
            current_row += 1

            for col_num, header_title in enumerate(report_columns, 1):
                cell = ws.cell(row=current_row, column=col_num, value=header_title)
                cell.font = header_font; cell.alignment = center_alignment; cell.border = border_thin; cell.fill = header_fill
            current_row += 1

            for idx, lead in enumerate(df_visit_done.to_dict('records'), 1):
                ws.cell(row=current_row, column=1, value=idx).alignment = center_alignment
                ws.cell(row=current_row, column=2, value=lead.get('bank_name', '')).alignment = left_alignment
                ws.cell(row=current_row, column=3, value=lead.get('branch_virtual', '')).alignment = left_alignment
                received_dt = lead.get('received_date'); date_str_recv = received_dt.strftime("%d-%b-%Y") if pd.notna(received_dt) and isinstance(received_dt, datetime.date) else ''
                ws.cell(row=current_row, column=4, value=date_str_recv).alignment = center_alignment
                ws.cell(row=current_row, column=5, value=lead.get('customer_name', '')).alignment = left_alignment
                ws.cell(row=current_row, column=6, value=lead.get('location', '')).alignment = left_alignment
                ws.cell(row=current_row, column=7, value=lead.get('site_engineer', '')).alignment = left_alignment
                deadline_dt = lead.get('deadline'); date_str_dead = deadline_dt.strftime("%d-%b-%Y") if pd.notna(deadline_dt) and isinstance(deadline_dt, datetime.date) else ''
                ws.cell(row=current_row, column=8, value=date_str_dead).alignment = center_alignment
                ws.cell(row=current_row, column=9, value=lead.get('status', '')).alignment = left_alignment

                for col_num in range(1, len(report_columns) + 1):
                    ws.cell(row=current_row, column=col_num).border = border_thin
                    ws.cell(row=current_row, column=col_num).font = cell_font
                current_row += 1

            current_row += 2 # Space before next section

            # VISIT PENDING Section
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(report_columns))
            title_cell_pending = ws.cell(row=current_row, column=1, value=f"VISIT PENDING - as on {current_date_str}")
            title_cell_pending.font = Font(name='Calibri', size=14, bold=True)
            title_cell_pending.alignment = center_alignment
            current_row += 1

            for col_num, header_title in enumerate(report_columns, 1):
                cell = ws.cell(row=current_row, column=col_num, value=header_title)
                cell.font = header_font; cell.alignment = center_alignment; cell.border = border_thin; cell.fill = header_fill
            current_row += 1

            for idx, lead in enumerate(df_visit_pending.to_dict('records'), 1):
                ws.cell(row=current_row, column=1, value=idx).alignment = center_alignment
                ws.cell(row=current_row, column=2, value=lead.get('bank_name', '')).alignment = left_alignment
                ws.cell(row=current_row, column=3, value=lead.get('branch_virtual', '')).alignment = left_alignment
                received_dt = lead.get('received_date'); date_str_recv = received_dt.strftime("%d-%b-%Y") if pd.notna(received_dt) and isinstance(received_dt, datetime.date) else ''
                ws.cell(row=current_row, column=4, value=date_str_recv).alignment = center_alignment
                ws.cell(row=current_row, column=5, value=lead.get('customer_name', '')).alignment = left_alignment
                ws.cell(row=current_row, column=6, value=lead.get('location', '')).alignment = left_alignment
                ws.cell(row=current_row, column=7, value=lead.get('site_engineer', 'N/A' if lead.get('status') == 'New' else lead.get('site_engineer', ''))).alignment = left_alignment
                deadline_dt = lead.get('deadline'); date_str_dead = deadline_dt.strftime("%d-%b-%Y") if pd.notna(deadline_dt) and isinstance(deadline_dt, datetime.date) else ''
                ws.cell(row=current_row, column=8, value=date_str_dead).alignment = center_alignment
                ws.cell(row=current_row, column=9, value=lead.get('status', '')).alignment = left_alignment

                for col_num in range(1, len(report_columns) + 1):
                    ws.cell(row=current_row, column=col_num).border = border_thin
                    ws.cell(row=current_row, column=col_num).font = cell_font
                current_row += 1

            custom_widths = [7, 30, 18, 15, 30, 40, 20, 15, 20]
            for i, col_letter in enumerate([get_column_letter(idx) for idx in range(1, len(report_columns) + 1)]):
                ws.column_dimensions[col_letter].width = custom_widths[i] if i < len(custom_widths) else 15

            excel_stream = BytesIO()
            wb.save(excel_stream)
            excel_stream.seek(0)
            return excel_stream

        # --- Main App UI Function ---
def build_mis_app():
    st.sidebar.header(f"Welcome, {st.session_state.get('username', 'Guest')}!")
    st.sidebar.write(f"Role: {st.session_state.get('role', 'N/A').upper()}")
    st.sidebar.markdown("---")
    st.sidebar.subheader("Filter Leads by Bank")
    if 'selected_bank_filter' not in st.session_state: st.session_state.selected_bank_filter = "-- All Banks --"
    current_filter_val = st.session_state.selected_bank_filter
    if current_filter_val not in ALL_BANK_OPTIONS_FILTER:
        st.session_state.selected_bank_filter = "-- All Banks --"
        current_filter_val = "-- All Banks --"
    selected_bank = st.sidebar.selectbox(
        "Select Bank:", options=ALL_BANK_OPTIONS_FILTER, key="bank_filter_nav_sidebar_v15",
        index=ALL_BANK_OPTIONS_FILTER.index(current_filter_val)
    )
    if selected_bank != st.session_state.selected_bank_filter:
        st.session_state.selected_bank_filter = selected_bank
        st.rerun()
    st.sidebar.markdown("---")
    if st.sidebar.button("Logout", key="logout_sidebar_button_main_v15"):
        keys_to_clear = ['logged_in', 'username', 'role', 'selected_bank_filter', 'daily_report_year', 'daily_report_month_name']
        for key in keys_to_clear:
            if key in st.session_state: del st.session_state[key]
        st.rerun()

    st.title(APP_NAME) # Uses globally defined APP_NAME
    top_cols = st.columns(4)
    with top_cols[0]:
        if st.button("🔄 Refresh Data", key="refresh_top_button_main_v15", help="Reload data from the database"): st.rerun()
    with top_cols[1]:
        if st.session_state.get('role') == 'admin':
            if st.button("📧 Check Emails", key="email_top_button_main_v15", help="Fetch new leads from the configured email account"):
                with st.spinner("Checking for new email leads..."): check_emails_once(); st.rerun()
    excel_dl_pl = top_cols[2].empty()
    custom_excel_dl_pl = top_cols[3].empty()

    db_list = run_db_query("SELECT * FROM office ORDER BY received_date DESC, id DESC", fetch_all=True)
    master_df = None
    if db_list:
        if PANDAS_AVAILABLE:
            try:
                master_df = pd.DataFrame(db_list)
                date_cols_to_convert = ['received_date', 'deadline', 'date_of_allocation',
                                        'visit_initiation_date', 'visit_completion_date', 'lead_completion_date']
                for col_name in date_cols_to_convert:
                    if col_name in master_df.columns:
                        master_df.loc[:, col_name] = pd.to_datetime(master_df[col_name], errors='coerce')
                        if pd.api.types.is_datetime64_any_dtype(master_df[col_name]) and master_df[col_name].dt.tz is not None:
                               master_df.loc[:, col_name] = master_df[col_name].dt.tz_localize(None)
            except Exception as e: st.error(f"Error converting database list to Pandas DataFrame: {e}"); master_df = None
        else: st.warning("Pandas library not available. Full data processing features might be limited.")

    if master_df is None and db_list: st.warning("Displaying raw list; Pandas DataFrame creation or processing failed.")

    filtered_df_display = master_df.copy() if master_df is not None else None
    list_for_display_fallback = db_list

    if filtered_df_display is not None and st.session_state.selected_bank_filter != "-- All Banks --":
        try:
            if 'bank_name' in filtered_df_display.columns:
                original_row_count = filtered_df_display.shape[0]
                filtered_df_display = filtered_df_display[filtered_df_display['bank_name'] == st.session_state.selected_bank_filter]
                if filtered_df_display.empty and original_row_count > 0:
                    list_for_display_fallback = [item for item in db_list if item.get('bank_name') == st.session_state.selected_bank_filter] if db_list else []
                elif not filtered_df_display.empty:
                    list_for_display_fallback = filtered_df_display.to_dict('records')
            else: st.warning("Cannot filter by bank: 'bank_name' column missing from data.")
        except Exception as e:
            st.error(f"Error during bank filtering: {e}")
            filtered_df_display = master_df.copy() if master_df is not None else None
            list_for_display_fallback = db_list

    if st.session_state.get('role') == 'admin':
        st.markdown("---"); st.header("Admin Dashboards & Reports")
        df_for_dashboards = filtered_df_display if filtered_df_display is not None else pd.DataFrame() # Ensure it's a DataFrame
        display_summary_dashboard_stats(df_for_dashboards)

        st.markdown("---")
        current_datetime_obj = datetime.datetime.now() # Renamed to avoid conflict
        available_years = list(range(current_datetime_obj.year - 3, current_datetime_obj.year + 2))
        month_names_list = [calendar.month_name[i] for i in range(1, 13)]

        if 'daily_report_year' not in st.session_state: st.session_state.daily_report_year = current_datetime_obj.year
        if 'daily_report_month_name' not in st.session_state: st.session_state.daily_report_month_name = month_names_list[current_datetime_obj.month - 1]

        dashboard_cols = st.columns([1,2])
        with dashboard_cols[0]:
            st.session_state.daily_report_year = st.selectbox(
                "Select Report Year:",
                available_years,
                index=available_years.index(st.session_state.daily_report_year) if st.session_state.daily_report_year in available_years else available_years.index(current_datetime_obj.year), # Robust index
                key="year_select_admin_dashboard_v15"
            )
        with dashboard_cols[1]:
            st.session_state.daily_report_month_name = st.selectbox(
                "Select Report Month:",
                month_names_list,
                index=month_names_list.index(st.session_state.daily_report_month_name) if st.session_state.daily_report_month_name in month_names_list else current_datetime_obj.month -1, # Robust index
                key="month_select_admin_dashboard_v15"
            )
        selected_month_number = month_names_list.index(st.session_state.daily_report_month_name) + 1

        display_per_day_allocation_dashboard(df_for_dashboards, st.session_state.daily_report_year, selected_month_number)

    st.markdown("---")
    st.header(f"Leads Details (Filter Applied: {st.session_state.selected_bank_filter})")

    active_data_for_display = filtered_df_display if PANDAS_AVAILABLE and filtered_df_display is not None else \
                            ([item for item in list_for_display_fallback if item.get('bank_name') == st.session_state.selected_bank_filter]
                                if st.session_state.selected_bank_filter != "-- All Banks --" and list_for_display_fallback else list_for_display_fallback)

    condition_met = False
    if active_data_for_display is not None:
        if PANDAS_AVAILABLE and isinstance(active_data_for_display, pd.DataFrame):
            if not active_data_for_display.empty:
                condition_met = True
        elif isinstance(active_data_for_display, list):
            if active_data_for_display:
                condition_met = True

    if condition_met:
        overdue_leads = []; due_soon_leads = []; on_hold_leads = []
        source_for_warnings = active_data_for_display.to_dict('records') if PANDAS_AVAILABLE and isinstance(active_data_for_display, pd.DataFrame) else active_data_for_display

        for lead_item in source_for_warnings:
            status_display_text, row_color = get_status_and_color_value(lead_item)
            lead_id_display = lead_item.get('id', 'N/A')
            lead_identifier = lead_item.get('bank_name', 'N/A')[:20]

            if "Overdue" in status_display_text: overdue_leads.append(f"🚨 ID {lead_id_display} ({lead_identifier})")
            elif row_color == '#ffdddd': due_soon_leads.append(f"⚠️ ID {lead_id_display} ({lead_identifier})")
            elif row_color == '#ffe8cc': on_hold_leads.append(f"🔔 ID {lead_id_display} ({lead_identifier})")

        if overdue_leads: st.error(f"**Overdue Leads:** {'; '.join(overdue_leads)}")
        if due_soon_leads: st.warning(f"**Leads Due Soon (Urgent):** {'; '.join(due_soon_leads)}")
        if on_hold_leads: st.info(f"**Leads On Hold / Nearing Deadline:** {'; '.join(on_hold_leads)}")

        if PANDAS_AVAILABLE and isinstance(active_data_for_display, pd.DataFrame) and not active_data_for_display.empty:
            df_to_style = active_data_for_display.copy()
            column_order_preference = [
                'id', 'bank_name', 'branch_virtual', 'customer_name', 'application_number',
                'location', 'contact_number', 'property_details', 'site_link',
                'received_date', 'date_of_allocation', 'deadline',
                'visit_initiation_date', 'visit_completion_date', 'lead_completion_date',
                'status', 'site_engineer', 'report_creator',
                'visit_type', 'distance', 'remarks', 'report_issue_notes',
                'admin_review_status', 'admin_comments', 'appraiser_quotation_obs'
                # 'site_photo_filenames', 'site_document_filenames' # Optionally add to main display
            ]
            actual_columns_to_display = [col for col in column_order_preference if col in df_to_style.columns]
            remaining_cols = [col for col in df_to_style.columns if col not in actual_columns_to_display]
            df_to_style_ordered = df_to_style[actual_columns_to_display + remaining_cols]

            def apply_row_styles(row_series_data):
                if row_series_data.name in df_to_style_ordered.index:
                    lead_dict_for_status = df_to_style_ordered.loc[row_series_data.name].to_dict()
                    _, color_hex = get_status_and_color_value(lead_dict_for_status)
                    return [f'background-color: {color_hex}'] * len(row_series_data)
                return [''] * len(row_series_data)


            for col_name_format in df_to_style_ordered.columns:
                if pd.api.types.is_datetime64_any_dtype(df_to_style_ordered[col_name_format]):
                    is_only_date = False
                    if not df_to_style_ordered[col_name_format].isna().all():
                        try:
                            valid_times = pd.to_datetime(df_to_style_ordered[col_name_format]).dropna().dt.time
                            if not valid_times.empty: # Ensure there are non-NaT values
                                is_only_date = (valid_times == datetime.time(0,0,0)).all()
                        except AttributeError: pass

                    if is_only_date:
                        df_to_style_ordered.loc[:, col_name_format] = pd.to_datetime(df_to_style_ordered[col_name_format], errors='coerce').dt.strftime('%Y-%m-%d')
                    else:
                        df_to_style_ordered.loc[:, col_name_format] = pd.to_datetime(df_to_style_ordered[col_name_format], errors='coerce').dt.strftime('%Y-%m-%d %H:%M')

            df_to_style_ordered = df_to_style_ordered.astype(str).replace({'None': '-', 'NaT': '-', 'nan':'-', 'nat':'-'})

            try:
                st.dataframe(df_to_style_ordered.style.apply(apply_row_styles, axis=1), use_container_width=True, height=450, hide_index=True)
            except Exception as e_style:
                st.error(f"Error applying styles to DataFrame: {e_style}")
                st.dataframe(df_to_style_ordered, use_container_width=True, height=450, hide_index=True)

        elif not PANDAS_AVAILABLE and active_data_for_display :
            st.warning("Pandas not available, displaying basic table.")
            st.table(active_data_for_display)

        if PANDAS_AVAILABLE and isinstance(active_data_for_display, pd.DataFrame) and not active_data_for_display.empty:
            df_for_standard_export = active_data_for_display.copy()
            for col in df_for_standard_export.select_dtypes(include=[np.datetime64, 'datetime64[ns]', 'datetime64[ns, UTC]']).columns: # Added UTC
                df_for_standard_export.loc[:, col] = pd.to_datetime(df_for_standard_export[col], errors='coerce').dt.strftime('%Y-%m-%d %H:%M:%S')
            df_for_standard_export = df_for_standard_export.fillna('')

            excel_filename_standard = f"MIS_Standard_Export_{st.session_state.selected_bank_filter.replace(' ','_')}_{datetime.date.today():%Y%m%d}.xlsx"
            try:
                output_stream_standard = BytesIO()
                with pd.ExcelWriter(output_stream_standard, engine='openpyxl') as excel_writer_std:
                    df_for_standard_export.to_excel(excel_writer_std, index=False, sheet_name='Filtered_Leads_Data')
                excel_dl_pl.download_button(
                    label="📄 Download Standard Excel",
                    data=output_stream_standard.getvalue(),
                    file_name=excel_filename_standard,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", # Corrected MIME
                    key="download_excel_standard_v15"
                )
            except Exception as e_std_excel: excel_dl_pl.error(f"Standard Excel DL Failed: {e_std_excel}")

            excel_stream_for_custom_report = generate_custom_excel_report(active_data_for_display)
            if excel_stream_for_custom_report:
                excel_filename_custom = f"MIS_Formatted_Report_{datetime.date.today():%d%b%Y}.xlsx"
                custom_excel_dl_pl.download_button(
                    label="📑 Download Custom MIS Report",
                    data=excel_stream_for_custom_report,
                    file_name=excel_filename_custom,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="download_excel_custom_v15"
                )
        elif not PANDAS_AVAILABLE:
            excel_dl_pl.warning("Pandas library needed for Excel downloads.")
            custom_excel_dl_pl.warning("Pandas library needed for Custom Excel report.")
    else: st.info(f"No leads data available to display for the current filter: '{st.session_state.selected_bank_filter}'.")

    if st.session_state.get('role') == 'admin':
        st.markdown("---"); st.subheader("Manually Add New Lead")
        with st.form("add_lead_form_admin_v15", clear_on_submit=True):
            form_col1, form_col2 = st.columns(2)
            with form_col1:
                selected_bank_form = st.selectbox("Bank Name*", options=ALL_BANK_OPTIONS_DROPDOWN, index=0, key="add_form_bank_select_v15")
                other_bank_name_form = ""
                if selected_bank_form == "Other": other_bank_name_form = st.text_input("Enter Other Bank Name*", key="add_form_bank_other_text_v15")
                customer_name_form = st.text_input("Customer Name", key="add_form_customer_name_v15")
                application_number_form = st.text_input("Application Number", key="add_form_app_number_v15")
                contact_number_form = st.text_input("Contact Number", key="add_form_contact_num_v15")
                distance_form = st.number_input("Distance (km)", min_value=0.0, step=0.1, format="%.1f", key="add_form_distance_v15", value=0.0) # Ensure value=0.0
                branch_virtual_form = st.text_input("Branch/Virtual Office", key="add_form_branch_virtual_v15")

            with form_col2:
                date_of_allocation_form = st.date_input("Date of Allocation (Optional)", value=None, format="YYYY-MM-DD", key="add_form_alloc_date_v15")
                deadline_date_form = st.date_input("Deadline Date (Optional)", value=None, format="YYYY-MM-DD", key="add_form_deadline_date_v15")
                location_form = st.text_area("Location Details", key="add_form_location_v15", height=68)
                site_link_form = st.text_input("Site Link (e.g., Google Maps)", key="add_form_site_link_v15")
                visit_type_form = st.text_input("Type of Visit", key="add_form_visit_type_v15")
                site_engineer_form = st.text_input("Site Engineer (Optional, if assigning now)", key="add_form_site_eng_v15")

            property_details_form = st.text_area("Property Details*", help="Detailed description of the property.", key="add_form_prop_details_v15")
            remarks_form = st.text_area("Initial Remarks (Optional)", key="add_form_remarks_v15")

            submitted_new_lead_form = st.form_submit_button("Save New Lead to Database")

            if submitted_new_lead_form:
                final_bank_name_to_save = other_bank_name_form.strip() if selected_bank_form == "Other" else selected_bank_form
                if final_bank_name_to_save == "--Select Bank--" : final_bank_name_to_save = ""

                if not final_bank_name_to_save or not property_details_form.strip():
                    st.error("Bank Name and Property Details are mandatory fields.")
                else:
                    new_lead_data_dict = {
                        'bank_name': final_bank_name_to_save,
                        'property_details': property_details_form.strip(),
                        'deadline': deadline_date_form if deadline_date_form else None,
                        'site_engineer': site_engineer_form.strip() or None,
                        'date_of_allocation': date_of_allocation_form if site_engineer_form.strip() and date_of_allocation_form else None,
                        'customer_name': customer_name_form.strip() or None,
                        'application_number': application_number_form.strip() or None,
                        'location': location_form.strip() or None,
                        'contact_number': contact_number_form.strip() or None,
                        'site_link': site_link_form.strip() or None,
                        'distance': distance_form if distance_form > 0 else None,
                        'visit_type': visit_type_form.strip() or None,
                        'remarks': remarks_form.strip() or None,
                        'branch_virtual': branch_virtual_form.strip() or None,
                        'status': 'Assigned Engineer' if site_engineer_form.strip() else 'New'
                    }
                    if add_lead_to_db(new_lead_data_dict):
                        st.success("New lead successfully added to the database!"); st.rerun()
                    else:
                        st.error("Failed to add the new lead. Check console logs for details.")

    st.markdown("---"); st.subheader("Perform Actions on a Selected Lead")
    if db_list:
        lead_action_options = {"": "--Select Lead ID--"}
        for lead_item_option in db_list:
            status_text_option, _ = get_status_and_color_value(lead_item_option)
            desc_text_option = lead_item_option.get('property_details', lead_item_option.get('bank_name', 'N/A'))[:30]
            lead_action_options[str(lead_item_option['id'])] = f"ID {lead_item_option['id']} - {desc_text_option}... ({status_text_option})"

        selected_lead_id_str = st.selectbox(
            "Select Lead for Action:",
            options=list(lead_action_options.keys()),
            format_func=lambda k: lead_action_options[k],
            key="action_lead_select_v15"
        )

        if selected_lead_id_str:
            selected_lead_details = next((item for item in db_list if str(item['id']) == selected_lead_id_str), None)
            if selected_lead_details:
                selected_lead_id_int = int(selected_lead_id_str)
                current_lead_status = selected_lead_details.get('status', 'New')

                st.markdown(f"--- \n #### Actions for Lead ID: {selected_lead_id_int} (Status: {current_lead_status})")
                # Define role-based booleans for clarity
                is_current_user_the_assigned_engineer = (
                    selected_lead_details.get('site_engineer') == st.session_state.get('username') and
                    st.session_state.get('role') == 'engineer'
                )
                is_admin = st.session_state.get('role') == 'admin'
                is_general_user = st.session_state.get('role') == 'user' # Make sure 'user' is the correct role string for general users
                is_report_creator = (
                    selected_lead_details.get('report_creator') == st.session_state.get('username') and
                    st.session_state.get('role') == 'user' # Assuming report creators might have 'user' role or a specific one
                ) # Adjust if report creators have a different role, e.g. 'creator'

                action_button_cols = st.columns(4)
                # ... (Existing action buttons: Assign Eng, Visit Done, Assign Creator, Report Done - keep their logic as is) ...
                with action_button_cols[0]: # Assign Engineer
                    can_assign_engineer = (current_lead_status == 'New' and is_admin)
                    if st.button("Assign Eng.", disabled=not can_assign_engineer, key=f"assign_eng_btn_{selected_lead_id_int}_v15", use_container_width=True, help="Assign a Site Engineer to this new lead."):
                        if can_assign_engineer:
                            st.session_state[f'show_assign_engineer_expander_{selected_lead_id_int}'] = not st.session_state.get(f'show_assign_engineer_expander_{selected_lead_id_int}', False)
                        else: st.warning("Only Admins can assign engineers to 'New' leads.")
                    if st.session_state.get(f'show_assign_engineer_expander_{selected_lead_id_int}', False) and can_assign_engineer:
                        with st.expander("Enter Engineer Name & Confirm", expanded=True):
                            engineer_name_input = st.text_input("Engineer Username:", key=f"assign_eng_name_txt_{selected_lead_id_int}_v15", value=selected_lead_details.get('site_engineer',''))
                            if st.button("Confirm Engineer Assignment", key=f"confirm_assign_eng_btn_{selected_lead_id_int}_v15"):
                                if engineer_name_input.strip():
                                    # ... (db update logic)
                                    update_query = "UPDATE office SET site_engineer=%s, status=%s, date_of_allocation=%s WHERE id=%s AND status=%s"
                                    params_tuple = (engineer_name_input.strip(), 'Assigned Engineer', datetime.date.today(), selected_lead_id_int, 'New')
                                    if run_db_query(update_query, params_tuple) is not None:
                                        st.success(f"Engineer '{engineer_name_input.strip()}' assigned successfully.")
                                        st.session_state[f'show_assign_engineer_expander_{selected_lead_id_int}']=False; st.rerun()
                                    else: st.error("Failed to assign engineer.")
                                else: st.warning("Engineer name cannot be empty.")
                
                with action_button_cols[1]: # Visit Done
                    can_mark_visit_done = (current_lead_status == 'Assigned Engineer' and (is_current_user_the_assigned_engineer or is_admin))
                    if st.button("✅ Visit Done", disabled=not can_mark_visit_done, key=f"mark_visit_done_btn_{selected_lead_id_int}_v15", use_container_width=True, help="Mark the site visit as completed."):
                        if can_mark_visit_done:
                            # ... (db update logic)
                            update_query = "UPDATE office SET status=%s, visit_completion_date=%s WHERE id=%s AND status=%s"
                            params_tuple = ('Visit Done', datetime.date.today(), selected_lead_id_int, 'Assigned Engineer')
                            if run_db_query(update_query, params_tuple) is not None: st.success("Site visit marked as done."); st.rerun()
                            else: st.error("Failed to mark visit as done.")
                        else: st.warning("Only the assigned Site Engineer or an Admin can mark the visit done.")

                with action_button_cols[2]: # Assign Creator
                    can_assign_report_creator = (current_lead_status == 'Visit Done' and not selected_lead_details.get('report_creator') and is_admin)
                    if st.button("Assign Creator", disabled=not can_assign_report_creator, key=f"assign_creator_btn_{selected_lead_id_int}_v15", use_container_width=True, help="Assign a Report Creator for this lead."):
                        if can_assign_report_creator:
                             st.session_state[f'show_assign_creator_expander_{selected_lead_id_int}'] = not st.session_state.get(f'show_assign_creator_expander_{selected_lead_id_int}', False)
                        else: st.warning("Only Admins can assign report creators to 'Visit Done' leads without one.")
                    if st.session_state.get(f'show_assign_creator_expander_{selected_lead_id_int}', False) and can_assign_report_creator:
                        with st.expander("Enter Report Creator Name & Confirm", expanded=True):
                            creator_name_input = st.text_input("Report Creator Username:", key=f"assign_creator_name_txt_{selected_lead_id_int}_v15", value=selected_lead_details.get('report_creator',''))
                            if st.button("Confirm Creator Assignment", key=f"confirm_assign_creator_btn_{selected_lead_id_int}_v15"):
                                if creator_name_input.strip():
                                    # ... (db update logic)
                                    update_query = "UPDATE office SET report_creator=%s, status=%s WHERE id=%s AND status=%s"
                                    params_tuple = (creator_name_input.strip(), 'Report in Progress', selected_lead_id_int, 'Visit Done')
                                    if run_db_query(update_query, params_tuple) is not None:
                                        st.success(f"Report Creator '{creator_name_input.strip()}' assigned.")
                                        st.session_state[f'show_assign_creator_expander_{selected_lead_id_int}']=False; st.rerun()
                                    else: st.error("Failed to assign report creator.")
                                else: st.warning("Creator name cannot be empty.")

                with action_button_cols[3]: # Report Done
                    # Determine if the current user is the assigned report creator
                    is_current_user_the_report_creator = (
                        selected_lead_details.get('report_creator') == st.session_state.get('username') and
                        selected_lead_details.get('report_creator') is not None # Ensure a creator is assigned
                    )
                    can_mark_report_done = (current_lead_status == 'Report in Progress' and (is_current_user_the_report_creator or is_admin))
                    if st.button("📝 Report Done", disabled=not can_mark_report_done, key=f"mark_report_done_btn_{selected_lead_id_int}_v15", use_container_width=True, help="Mark the report as completed and ready for admin review."):
                        if can_mark_report_done:
                            # ... (db update logic)
                            update_query = "UPDATE office SET status=%s, admin_review_status=%s, lead_completion_date=%s WHERE id=%s AND status=%s"
                            params_tuple = ('Completed', 'Pending Review', datetime.date.today(), selected_lead_id_int, 'Report in Progress')
                            if run_db_query(update_query, params_tuple) is not None: st.success("Report marked as done."); st.rerun()
                            else: st.error("Failed to mark report as done.")
                        else: st.warning("Only the assigned Report Creator or an Admin can mark the report done.")


                # ... (Existing "Update Details for Lead ID", "Report/Issue Notes", "Admin Review" sections - keep their logic as is) ...
                if current_lead_status != 'Completed' or is_admin:
                    st.markdown("---"); st.subheader(f"Update Details for Lead ID {selected_lead_id_int}")
                    # ... (your existing update details form and logic)
                    update_col1, update_col2 = st.columns(2)
                    with update_col1:
                        contact_update = st.text_input("Contact Number:", value=selected_lead_details.get('contact_number',''), key=f"update_contact_{selected_lead_id_int}_v15")
                        visit_type_update = st.text_input("Visit Type:", value=selected_lead_details.get('visit_type',''), key=f"update_visit_type_{selected_lead_id_int}_v15")
                        branch_virtual_update = st.text_input("Branch/Virtual:", value=selected_lead_details.get('branch_virtual',''), key=f"update_branch_{selected_lead_id_int}_v15")
                    with update_col2:
                        location_update = st.text_area("Location:", value=selected_lead_details.get('location',''), key=f"update_location_{selected_lead_id_int}_v15", height=100)
                    remarks_update = st.text_area("Remarks:", value=selected_lead_details.get('remarks',''), key=f"update_remarks_{selected_lead_id_int}_v15", height=100)
                    if st.button("Save Updates to Details", key=f"save_details_update_btn_{selected_lead_id_int}_v15"):
                        fields_to_update = {}
                        # ... (logic to populate fields_to_update)
                        if contact_update != (selected_lead_details.get('contact_number','') or ''): fields_to_update['contact_number'] = contact_update.strip() or None
                        if visit_type_update != (selected_lead_details.get('visit_type','') or ''): fields_to_update['visit_type'] = visit_type_update.strip() or None
                        if branch_virtual_update != (selected_lead_details.get('branch_virtual','') or ''): fields_to_update['branch_virtual'] = branch_virtual_update.strip() or None
                        if location_update != (selected_lead_details.get('location','') or ''): fields_to_update['location'] = location_update.strip() or None
                        if remarks_update != (selected_lead_details.get('remarks','') or ''): fields_to_update['remarks'] = remarks_update.strip() or None
                        if fields_to_update:
                            # ... (db update logic)
                            set_clause_parts = [f"`{col_name}`=%s" for col_name in fields_to_update.keys()]
                            update_query_details = f"UPDATE office SET {', '.join(set_clause_parts)} WHERE id=%s"
                            update_params_details = list(fields_to_update.values()) + [selected_lead_id_int]
                            if run_db_query(update_query_details, tuple(update_params_details)) is not None:
                                st.success("Lead details updated successfully."); st.rerun()
                            else: st.error("Failed to update lead details.")
                        else: st.info("No changes detected in the details to save.")
                
                can_edit_notes = (
                    is_admin or 
                    is_current_user_the_assigned_engineer or 
                    (selected_lead_details.get('report_creator') == st.session_state.get('username') and selected_lead_details.get('report_creator'))
                )
                if current_lead_status in ['Assigned Engineer', 'Visit Done', 'Report in Progress','Completed'] and can_edit_notes:
                    st.markdown("---"); st.subheader(f"Report/Issue Notes for Lead ID {selected_lead_id_int}")
                    # ... (your existing notes update form and logic)
                    current_notes_value = selected_lead_details.get('report_issue_notes','') or ''
                    updated_notes_value = st.text_area("Notes:", value=current_notes_value, key=f"update_notes_txt_{selected_lead_id_int}_v15", height=100)
                    if st.button("Save Notes", key=f"save_notes_btn_{selected_lead_id_int}_v15"):
                        if current_notes_value != updated_notes_value:
                            query_save_notes = "UPDATE office SET report_issue_notes=%s WHERE id=%s"
                            if run_db_query(query_save_notes, (updated_notes_value.strip() or None, selected_lead_id_int)) is not None:
                                st.success("Notes saved successfully."); st.rerun()
                            else: st.error("Failed to save notes.")
                        else: st.info("No changes detected in notes to save.")

                if is_admin and current_lead_status == 'Completed':
                    st.markdown("---"); st.subheader(f"Admin Review for Lead ID {selected_lead_id_int}")
                    # ... (your existing admin review form and logic)
                    current_review_status = selected_lead_details.get('admin_review_status','Pending Review') or 'Pending Review'
                    st.write(f"Current Review Status: **{current_review_status}**")
                    report_creator_notes = selected_lead_details.get('report_issue_notes','');
                    if report_creator_notes: st.info(f"Report Creator's Notes:\n{report_creator_notes}")
                    review_status_options = ['Pending Review','Approved','Rejected - Needs Revision']
                    current_option_index = review_status_options.index(current_review_status) if current_review_status in review_status_options else 0
                    new_review_status_selection = st.selectbox("Set Review Status:", review_status_options, index=current_option_index, key=f"admin_review_status_select_{selected_lead_id_int}_v15")
                    admin_comments_input = st.text_area("Admin Comments:", key=f"admin_review_comments_txt_{selected_lead_id_int}_v15", value=selected_lead_details.get('admin_comments',''))
                    if st.button("Save Admin Review", key=f"save_admin_review_btn_{selected_lead_id_int}_v15"):
                        new_overall_status_for_lead = current_lead_status
                        if new_review_status_selection == 'Rejected - Needs Revision':
                            new_overall_status_for_lead = 'Report in Progress'
                        query_admin_review = "UPDATE office SET admin_review_status=%s, status=%s, admin_comments=%s WHERE id=%s"
                        params_admin_review = (new_review_status_selection, new_overall_status_for_lead, admin_comments_input.strip() or None, selected_lead_id_int)
                        if run_db_query(query_admin_review, params_admin_review) is not None:
                            st.success("Admin review saved successfully."); st.rerun()
                        else: st.error("Failed to save admin review.")


                # --- BEGIN: Modified File Upload Section (Site Engineer Only) ---
                is_appropriate_status_for_upload = current_lead_status in ['Assigned Engineer', 'Visit Done', 'Report in Progress', 'Completed']
                if is_current_user_the_assigned_engineer and is_appropriate_status_for_upload:
                    st.markdown("---")
                    st.subheader(f"Upload Site Photos & Documents for Lead ID {selected_lead_id_int}")
                    st.caption("Only the assigned site engineer can upload files.")

                    uploaded_photos = st.file_uploader(
                        "Upload New Site Photos (e.g., JPG, PNG)",
                        type=["png", "jpg", "jpeg"],
                        accept_multiple_files=True,
                        key=f"photo_uploader_{selected_lead_id_int}_v15_eng" # Unique key for engineer
                    )
                    uploaded_docs = st.file_uploader(
                        "Upload New Site Documents (e.g., PDF, DOCX)",
                        type=["pdf", "doc", "docx", "xls", "xlsx", "txt"],
                        accept_multiple_files=True,
                        key=f"doc_uploader_{selected_lead_id_int}_v15_eng" # Unique key for engineer
                    )

                    if st.button("Process Uploaded Files", key=f"save_files_btn_{selected_lead_id_int}_v15_eng"):
                        LEAD_FILES_BASE_PATH = os.path.join(APP_ROOT_PATH, "instance", "lead_uploads")
                        os.makedirs(LEAD_FILES_BASE_PATH, exist_ok=True)
                        lead_specific_upload_dir = os.path.join(LEAD_FILES_BASE_PATH, str(selected_lead_id_int))
                        os.makedirs(lead_specific_upload_dir, exist_ok=True)

                        new_photos_saved_this_session = []
                        if uploaded_photos:
                            photos_dir = os.path.join(lead_specific_upload_dir, "photos")
                            os.makedirs(photos_dir, exist_ok=True)
                            for photo in uploaded_photos:
                                file_path = os.path.join(photos_dir, photo.name)
                                with open(file_path, "wb") as f: f.write(photo.getbuffer())
                                new_photos_saved_this_session.append(photo.name)
                            if new_photos_saved_this_session:
                                st.toast(f"{len(new_photos_saved_this_session)} photo(s) saved to server.", icon="📤")

                        new_docs_saved_this_session = []
                        if uploaded_docs:
                            docs_dir = os.path.join(lead_specific_upload_dir, "documents")
                            os.makedirs(docs_dir, exist_ok=True)
                            for doc in uploaded_docs:
                                file_path = os.path.join(docs_dir, doc.name)
                                with open(file_path, "wb") as f: f.write(doc.getbuffer())
                                new_docs_saved_this_session.append(doc.name)
                            if new_docs_saved_this_session:
                                st.toast(f"{len(new_docs_saved_this_session)} document(s) saved to server.", icon="📤")

                        if new_photos_saved_this_session or new_docs_saved_this_session:
                            current_files_data_db = run_db_query(
                                "SELECT site_photo_filenames, site_document_filenames FROM office WHERE id = %s",
                                (selected_lead_id_int,), fetch_one=True
                            )
                            db_photo_list = []
                            if current_files_data_db and current_files_data_db.get('site_photo_filenames'):
                                try: db_photo_list = json.loads(current_files_data_db['site_photo_filenames'])
                                except (json.JSONDecodeError, TypeError): db_photo_list = []
                            db_doc_list = []
                            if current_files_data_db and current_files_data_db.get('site_document_filenames'):
                                try: db_doc_list = json.loads(current_files_data_db['site_document_filenames'])
                                except (json.JSONDecodeError, TypeError): db_doc_list = []
                            
                            updated_photo_list = sorted(list(set(db_photo_list + new_photos_saved_this_session)))
                            updated_doc_list = sorted(list(set(db_doc_list + new_docs_saved_this_session)))

                            update_files_query_db = "UPDATE office SET site_photo_filenames=%s, site_document_filenames=%s WHERE id=%s"
                            params_files_db = (
                                json.dumps(updated_photo_list) if updated_photo_list else None,
                                json.dumps(updated_doc_list) if updated_doc_list else None,
                                selected_lead_id_int
                            )
                            if run_db_query(update_files_query_db, params_files_db) is not None:
                                st.success("File references updated in DB."); st.rerun()
                            else: st.error("Failed to update file references in DB.")
                        elif not uploaded_photos and not uploaded_docs:
                            st.info("No new files were selected for upload.")
                # --- END: Modified File Upload Section ---

                # --- BEGIN: View & Download Site Files Section (Admin, User, Assigned Engineer) ---
                can_view_or_download_files = is_admin or is_general_user or is_current_user_the_assigned_engineer
                
                if can_view_or_download_files:
                    files_data = run_db_query(
                        "SELECT site_photo_filenames, site_document_filenames FROM office WHERE id = %s",
                        (selected_lead_id_int,), fetch_one=True
                    )

                    if files_data and (files_data.get('site_photo_filenames') or files_data.get('site_document_filenames')):
                        st.markdown("---")
                        st.subheader(f"View/Download Site Files for Lead ID {selected_lead_id_int}")
                        
                        LEAD_FILES_VIEW_BASE_PATH = os.path.join(APP_ROOT_PATH, "instance", "lead_uploads")

                        # Display Photos with Download Buttons
                        db_photos_json = files_data.get('site_photo_filenames')
                        photo_filenames = []
                        if db_photos_json:
                            try: photo_filenames = json.loads(db_photos_json)
                            except (json.JSONDecodeError, TypeError): pass
                        
                        if photo_filenames:
                            st.markdown("**Photos:**")
                            for photo_name in photo_filenames:
                                photo_file_path = os.path.join(LEAD_FILES_VIEW_BASE_PATH, str(selected_lead_id_int), "photos", photo_name)
                                file_bytes_content = None
                                if os.path.exists(photo_file_path):
                                    try:
                                        with open(photo_file_path, "rb") as pf: file_bytes_content = pf.read()
                                    except Exception as e_read: print(f"Error reading photo {photo_file_path}: {e_read}")
                                
                                if file_bytes_content:
                                    mime_type = "image/jpeg" 
                                    if photo_name.lower().endswith(".png"): mime_type = "image/png"
                                    elif photo_name.lower().endswith(".jpg"): mime_type = "image/jpeg"
                                    st.download_button(label=f"⬇️ {photo_name}", data=file_bytes_content, file_name=photo_name, mime=mime_type, key=f"dl_photo_{selected_lead_id_int}_{photo_name}")
                                else: st.caption(f"📄 {photo_name} (File not found/readable on server)")
                        elif db_photos_json is not None : # Field exists but might be empty JSON array "[]" or invalid
                            st.caption("No photos found or photo list is corrupted.")
                        # else: # No photo filenames recorded at all (field is NULL or doesn't exist, though we assume it exists)
                            # st.caption("No photos recorded for this lead.")


                        # Display Documents with Download Buttons
                        db_docs_json = files_data.get('site_document_filenames')
                        doc_filenames = []
                        if db_docs_json:
                            try: doc_filenames = json.loads(db_docs_json)
                            except (json.JSONDecodeError, TypeError): pass

                        if doc_filenames:
                            st.markdown("**Documents:**")
                            for doc_name in doc_filenames:
                                doc_file_path = os.path.join(LEAD_FILES_VIEW_BASE_PATH, str(selected_lead_id_int), "documents", doc_name)
                                file_bytes_content = None
                                if os.path.exists(doc_file_path):
                                    try:
                                        with open(doc_file_path, "rb") as df_file: file_bytes_content = df_file.read() # Renamed df to df_file
                                    except Exception as e_read: print(f"Error reading doc {doc_file_path}: {e_read}")

                                if file_bytes_content:
                                    st.download_button(label=f"⬇️ {doc_name}",data=file_bytes_content,file_name=doc_name,mime="application/octet-stream",key=f"dl_doc_{selected_lead_id_int}_{doc_name}")
                                else: st.caption(f"📄 {doc_name} (File not found/readable on server)")
                        elif db_docs_json is not None:
                             st.caption("No documents found or document list is corrupted.")
                        # else:
                            # st.caption("No documents recorded for this lead.")
                    # else: # This case means either files_data is None OR both filename fields are None/empty.
                         # Handled by the individual "No photos/documents found" messages if lists are empty.
                         # st.caption("No site files are currently associated with this lead.") # Could be added if no files_data at all
                # --- END: View & Download Site Files Section ---

            else: st.warning("Selected lead ID data could not be found. Please refresh.")
        else: st.info("Select a Lead ID from the dropdown above to view actions and details.")
    else: st.info("No leads available in the database to perform actions on.")

# --- App Entry Point Logic ---
# (Your existing app entry point logic remains here: session state init, db_ok check, login screen, or build_mis_app call)
# ...
# Ensure this part is identical to the previous version
if 'logged_in' not in st.session_state: st.session_state['logged_in'] = False
if 'username' not in st.session_state: st.session_state['username'] = None
if 'role' not in st.session_state: st.session_state['role'] = None

if 'db_ok' not in st.session_state:
    st.session_state['db_ok'] = initialize_database()

if not st.session_state.get('db_ok', False):
    st.error("CRITICAL ERROR: Database connection failed or required tables not found. Check logs and settings. Ensure 'admins', 'users', 'office', and 'site_engineers' tables exist in your MySQL database. \n\n For file upload functionality, also ensure 'office' table has `site_photo_filenames` (TEXT) and `site_document_filenames` (TEXT) columns.")
elif not st.session_state.get('logged_in', False):
    login_col1, login_col2, login_col3 = st.columns([1, 1.5, 1])
    with login_col2:
        st.title(f"{APP_NAME} Portal")
        admin_tab, user_tab, engineer_tab = st.tabs(["Admin Login/Signup", "User Login/Signup", "Site Engineer Login/Signup"])
        # ... (Rest of your login/signup tabs logic remains here) ...
        with admin_tab:
            admin_action_choice = st.radio("Admin Action:", ("Sign In", "Sign Up"), horizontal=True, key="admin_action_radio_v15")
            if admin_action_choice == "Sign In":
                st.subheader("Admin Sign In")
                with st.form("admin_login_form_v15"):
                    admin_username = st.text_input("Admin Username", key="admin_uname_login_input_v15")
                    admin_password = st.text_input("Admin Password", type="password", key="admin_pword_login_input_v15")
                    submitted_admin_login = st.form_submit_button("Sign In as Admin")
                    if submitted_admin_login:
                        if not admin_username or not admin_password: st.error("Username and Password are required.")
                        else:
                            q = "SELECT admin_id, username, password_hash FROM admins WHERE username = %s"
                            data = run_db_query(q, (admin_username,), fetch_one=True)
                            if data and verify_password(admin_password, data['password_hash']):
                                st.session_state['logged_in'] = True; st.session_state['username'] = data['username']; st.session_state['role'] = 'admin'; st.rerun()
                            else: st.error("Invalid admin username or password.")
            elif admin_action_choice == "Sign Up":
                st.subheader("Admin Sign Up (Restricted)")
                with st.form("admin_signup_form_v15"):
                    su_admin_user = st.text_input("Choose Admin Username*", key="admin_uname_signup_input_v15")
                    su_admin_pass = st.text_input("Choose Admin Password*", type="password", key="admin_pword_signup_input_v15")
                    su_admin_conf = st.text_input("Confirm Admin Password*", type="password", key="admin_pword_conf_signup_input_v15")
                    su_admin_secr = st.text_input("Admin Secret Code*", type="password", key="admin_secret_signup_input_v15")
                    submitted_admin_signup = st.form_submit_button("Sign Up as Admin")
                    if submitted_admin_signup:
                        if not all([su_admin_user, su_admin_pass, su_admin_conf, su_admin_secr]): st.error("All fields are required for admin signup.")
                        elif su_admin_pass != su_admin_conf: st.error("Passwords do not match.")
                        elif su_admin_secr != ADMIN_SIGNUP_SECRET: st.error("Incorrect Admin Secret Code.")
                        else:
                            q_ex_admin = """
                                SELECT username FROM users WHERE username = %s
                                UNION SELECT username FROM admins WHERE username = %s
                                UNION SELECT username FROM site_engineers WHERE username = %s
                            """
                            ex_u_admin = run_db_query(q_ex_admin, (su_admin_user, su_admin_user, su_admin_user), fetch_one=True)
                            if ex_u_admin: st.error("This username is already taken.")
                            else:
                                h_pw_admin = get_password_hash(su_admin_pass)
                                if h_pw_admin:
                                    q_ins_admin = "INSERT INTO admins (username, password_hash) VALUES (%s, %s)"
                                    if run_db_query(q_ins_admin, (su_admin_user, h_pw_admin)) is not None: st.success("Admin account created! Please Sign In.")
                                    else: st.error("Admin signup failed. Please try again.")
                                else: st.error("Password hashing failed. Admin account cannot be created.")
        with user_tab:
            user_action_choice = st.radio("User Action:", ("Sign In", "Sign Up"), horizontal=True, key="user_action_radio_v15")
            if user_action_choice == "Sign In":
                st.subheader("User Sign In")
                with st.form("user_login_form_v15"):
                    user_username = st.text_input("Username", key="user_uname_login_input_v15")
                    user_password = st.text_input("Password", type="password", key="user_pword_login_input_v15")
                    submitted_user_login = st.form_submit_button("Sign In")
                    if submitted_user_login:
                        if not user_username or not user_password: st.error("Username and Password are required.")
                        else:
                            q_user = "SELECT user_id, username, password_hash, role FROM users WHERE username = %s" # Ensure your 'users' table has a 'role' column
                            data_user = run_db_query(q_user, (user_username,), fetch_one=True)
                            if data_user and verify_password(user_password, data_user['password_hash']):
                                st.session_state['logged_in'] = True
                                st.session_state['username'] = data_user['username']
                                st.session_state['role'] = data_user.get('role', 'user') # Default to 'user' if role column is missing or NULL
                                st.rerun()
                            else: st.error("Invalid username or password.")
            elif user_action_choice == "Sign Up":
                st.subheader("User Sign Up")
                with st.form("user_signup_form_v15"):
                    su_gen_user = st.text_input("Choose Username*", key="user_uname_signup_input_v15")
                    su_gen_pass = st.text_input("Choose Password*", type="password", key="user_pword_signup_input_v15")
                    su_gen_conf = st.text_input("Confirm Password*", type="password", key="user_pword_conf_signup_input_v15")
                    # Assuming 'user' role is default for this signup. If you want to assign roles, add a field.
                    submitted_user_signup = st.form_submit_button("Sign Up")
                    if submitted_user_signup:
                        if not all([su_gen_user, su_gen_pass, su_gen_conf]): st.error("All fields are required for user signup.")
                        elif su_gen_pass != su_gen_conf: st.error("Passwords do not match.")
                        else:
                            q_ex_user = """
                                SELECT username FROM users WHERE username = %s
                                UNION SELECT username FROM admins WHERE username = %s
                                UNION SELECT username FROM site_engineers WHERE username = %s
                            """
                            ex_u_user = run_db_query(q_ex_user, (su_gen_user, su_gen_user, su_gen_user), fetch_one=True)
                            if ex_u_user: st.error("This username is already taken.")
                            else:
                                h_pw_user = get_password_hash(su_gen_pass)
                                if h_pw_user:
                                    # Ensure your 'users' table can accept role or has a default, e.g., 'user'
                                    q_ins_user = "INSERT INTO users (username, password_hash, role) VALUES (%s, %s, %s)"
                                    if run_db_query(q_ins_user, (su_gen_user, h_pw_user, 'user')) is not None: st.success("User account created! Please Sign In.")
                                    else: st.error("User signup failed. Please try again.")
                                else: st.error("Password hashing failed. User account cannot be created.")
        
        with engineer_tab:
            engineer_action_choice = st.radio("Site Engineer Action:", ("Sign In", "Sign Up"), horizontal=True, key="engineer_action_radio_v15")
            if engineer_action_choice == "Sign In":
                st.subheader("Site Engineer Sign In")
                with st.form("engineer_login_form_v15"):
                    eng_username_login = st.text_input("Site Engineer Username", key="eng_uname_login_input_v15")
                    eng_password_login = st.text_input("Site Engineer Password", type="password", key="eng_pword_login_input_v15")
                    submitted_eng_login = st.form_submit_button("Sign In as Site Engineer")
                    if submitted_eng_login:
                        if not eng_username_login or not eng_password_login: st.error("Username and Password are required.")
                        else:
                            q_eng = "SELECT engineer_id, username, password_hash FROM site_engineers WHERE username = %s"
                            data_eng = run_db_query(q_eng, (eng_username_login,), fetch_one=True)
                            if data_eng and verify_password(eng_password_login, data_eng['password_hash']):
                                st.session_state['logged_in'] = True
                                st.session_state['username'] = data_eng['username']
                                st.session_state['role'] = 'engineer'
                                st.rerun()
                            else: st.error("Invalid Site Engineer username or password.")
            elif engineer_action_choice == "Sign Up":
                st.subheader("Site Engineer Sign Up")
                with st.form("engineer_signup_form_v15"):
                    su_site_eng_user = st.text_input("Choose Site Engineer Username*", key="eng_uname_signup_input_v15")
                    su_site_eng_pass = st.text_input("Choose Site Engineer Password*", type="password", key="eng_pword_signup_input_v15")
                    su_site_eng__conf = st.text_input("Confirm Site Engineer Password*", type="password", key="eng_pword_conf_signup_input_v15") # Corrected variable name
                    submitted_eng_signup = st.form_submit_button("Sign Up as Site Engineer")
                    if submitted_eng_signup:
                        if not all([su_site_eng_user, su_site_eng_pass, su_site_eng_conf]): # Use corrected su_site_eng_conf
                            st.error("Username and Password fields are required for signup.")
                        elif su_site_eng_pass != su_site_eng_conf: # Use corrected su_site_eng_conf
                            st.error("Passwords do not match.")
                        else:
                            q_ex_eng = """
                                SELECT username FROM users WHERE username = %s
                                UNION SELECT username FROM admins WHERE username = %s
                                UNION SELECT username FROM site_engineers WHERE username = %s
                            """
                            ex_u_eng = run_db_query(q_ex_eng, (su_site_eng_user, su_site_eng_user, su_site_eng_user), fetch_one=True)
                            if ex_u_eng: st.error("This username is already taken.")
                            else:
                                h_pw_eng = get_password_hash(su_site_eng_pass)
                                if h_pw_eng:
                                    q_ins_eng = "INSERT INTO site_engineers (username, password_hash) VALUES (%s, %s)"
                                    if run_db_query(q_ins_eng, (su_site_eng_user, h_pw_eng)) is not None:
                                        st.success("Site Engineer account created! Please Sign In.")
                                    else: st.error("Site Engineer signup failed. Please try again.")
                                else: st.error("Password hashing failed. Account cannot be created.")
else:
    build_mis_app()